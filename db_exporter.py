#MIT
import csv
import sqlite3
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.utils import get_column_letter
import os
import sys
import threading
import time

class DBExporter:
    def __init__(self, root, default_format="xlsx"):
        self.root = root
        self.root.title("数据库表导出工具")
        self.root.geometry("700x600")
        self.root.resizable(True, True)
        
        # 变量初始化
        self.db_path = ""
        self.selected_tables = []
        self.export_path = ""
        self.export_filename = "export.xlsx"
        self.export_format = default_format if default_format in ("xlsx", "csv") else "xlsx"
        self.is_exporting = False
        self.export_cancelled = False
        self.log_messages = []
        
        # 创建主框架
        self.main_frame = ttk.Frame(root, padding="10")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 创建数据库选择区域
        self.create_db_selection_frame()
        
        # 创建数据表选择区域
        self.create_table_selection_frame()
        
        # 创建导出配置区域
        self.create_export_config_frame()
        
        # 创建操作按钮区域
        self.create_button_frame()
        
        # 创建进度显示区域
        self.create_progress_frame()
        
        # 创建日志显示区域
        self.create_log_frame()
    
    def create_db_selection_frame(self):
        """创建数据库选择区域"""
        frame = ttk.LabelFrame(self.main_frame, text="数据库选择", padding="10")
        frame.pack(fill=tk.X, pady=5)
        
        self.db_path_var = tk.StringVar()
        
        ttk.Label(frame, text="数据库文件:").pack(side=tk.LEFT, padx=5)
        ttk.Entry(frame, textvariable=self.db_path_var, width=50).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Button(frame, text="浏览", command=self.browse_db).pack(side=tk.LEFT, padx=5)
        ttk.Button(frame, text="连接", command=self.connect_db).pack(side=tk.LEFT, padx=5)
    
    def create_table_selection_frame(self):
        """创建数据表选择区域"""
        frame = ttk.LabelFrame(self.main_frame, text="数据表选择", padding="10")
        frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # 全选/取消全选按钮
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=tk.X, pady=5)
        ttk.Button(btn_frame, text="全选", command=self.select_all_tables).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="取消全选", command=self.deselect_all_tables).pack(side=tk.LEFT, padx=5)
        
        # 表列表
        self.table_list_frame = ttk.Frame(frame)
        self.table_list_frame.pack(fill=tk.BOTH, expand=True)
        
        self.table_list_scrollbar = ttk.Scrollbar(self.table_list_frame, orient=tk.VERTICAL)
        self.table_list = tk.Listbox(self.table_list_frame, selectmode=tk.MULTIPLE, yscrollcommand=self.table_list_scrollbar.set)
        self.table_list_scrollbar.config(command=self.table_list.yview)
        self.table_list_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.table_list.pack(fill=tk.BOTH, expand=True)
    
    def create_export_config_frame(self):
        """创建导出配置区域"""
        frame = ttk.LabelFrame(self.main_frame, text="导出配置", padding="10")
        frame.pack(fill=tk.X, pady=5)
        
        # 文件路径选择
        path_frame = ttk.Frame(frame)
        path_frame.pack(fill=tk.X, pady=5)
        ttk.Label(path_frame, text="保存路径:").pack(side=tk.LEFT, padx=5)
        self.export_path_var = tk.StringVar(value=os.getcwd())
        ttk.Entry(path_frame, textvariable=self.export_path_var, width=50).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Button(path_frame, text="浏览", command=self.browse_export_path).pack(side=tk.LEFT, padx=5)
        
        # 文件名设置
        filename_frame = ttk.Frame(frame)
        filename_frame.pack(fill=tk.X, pady=5)
        ttk.Label(filename_frame, text="文件名:").pack(side=tk.LEFT, padx=5)
        self.export_filename_var = tk.StringVar(value="export")
        ttk.Entry(filename_frame, textvariable=self.export_filename_var, width=30).pack(side=tk.LEFT, padx=5)
        self.export_suffix_label = ttk.Label(filename_frame, text=".xlsx")
        self.export_suffix_label.pack(side=tk.LEFT)

        format_frame = ttk.Frame(frame)
        format_frame.pack(fill=tk.X, pady=5)
        ttk.Label(format_frame, text="导出格式:").pack(side=tk.LEFT, padx=5)
        self.export_format_var = tk.StringVar(value=self.export_format)
        format_combo = ttk.Combobox(
            format_frame,
            textvariable=self.export_format_var,
            state="readonly",
            values=["xlsx", "csv"],
            width=10,
        )
        format_combo.pack(side=tk.LEFT, padx=5)
        format_combo.bind("<<ComboboxSelected>>", lambda _e: self.on_format_change())
        self.format_hint_label = ttk.Label(format_frame, text="")
        self.format_hint_label.pack(side=tk.LEFT, padx=8)
        self.on_format_change()
    
    def create_button_frame(self):
        """创建操作按钮区域"""
        frame = ttk.Frame(self.main_frame)
        frame.pack(fill=tk.X, pady=5)
        
        self.export_button = ttk.Button(frame, text="导出", command=self.start_export)
        self.export_button.pack(side=tk.RIGHT, padx=5)
        self.cancel_button = ttk.Button(frame, text="取消", command=self.cancel_export, state=tk.DISABLED)
        self.cancel_button.pack(side=tk.RIGHT, padx=5)
    
    def create_progress_frame(self):
        """创建进度显示区域"""
        frame = ttk.LabelFrame(self.main_frame, text="导出进度", padding="10")
        frame.pack(fill=tk.X, pady=5)
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, pady=5)
        
        self.progress_label = ttk.Label(frame, text="就绪")
        self.progress_label.pack(side=tk.LEFT, padx=5)
    
    def create_log_frame(self):
        """创建日志显示区域"""
        frame = ttk.LabelFrame(self.main_frame, text="操作日志", padding="10")
        frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.log_text = tk.Text(frame, height=10, state=tk.DISABLED)
        self.log_scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.config(yscrollcommand=self.log_scrollbar.set)
        self.log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.pack(fill=tk.BOTH, expand=True)
    
    def browse_db(self):
        """浏览选择数据库文件"""
        file_path = filedialog.askopenfilename(
            title="选择SQLite3数据库文件",
            filetypes=[("SQLite3数据库文件", "*.db *.sqlite *.sqlite3"), ("所有文件", "*.*")]
        )
        if file_path:
            self.db_path_var.set(file_path)
    
    def browse_export_path(self):
        """浏览选择导出路径"""
        directory = filedialog.askdirectory(title="选择导出路径")
        if directory:
            self.export_path_var.set(directory)

    def on_format_change(self):
        """更新导出格式相关文案"""
        self.export_format = self.export_format_var.get() or "xlsx"
        if self.export_format == "csv":
            self.export_suffix_label.config(text=".csv")
            self.format_hint_label.config(text="CSV模式将按“每表一个文件”导出，文件名与表名相同")
        else:
            self.export_suffix_label.config(text=".xlsx")
            self.format_hint_label.config(text="XLSX模式将导出为一个工作簿，所选表各自独立工作表")
    
    def connect_db(self):
        """连接到数据库并获取表列表"""
        self.db_path = self.db_path_var.get()
        if not self.db_path:
            messagebox.showerror("错误", "请选择数据库文件")
            return
        
        if not os.path.exists(self.db_path):
            messagebox.showerror("错误", "数据库文件不存在")
            return
        
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # 获取所有表名
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = cursor.fetchall()
            
            # 清空表列表
            self.table_list.delete(0, tk.END)
            
            # 添加表到列表
            for table in tables:
                table_name = table[0]
                self.table_list.insert(tk.END, table_name)
            
            conn.close()
            self.log("成功连接到数据库并获取表列表")
            messagebox.showinfo("成功", f"成功连接到数据库，发现 {len(tables)} 个表")
        except Exception as e:
            self.log(f"连接数据库失败: {e}")
            messagebox.showerror("错误", f"连接数据库失败: {e}")
    
    def select_all_tables(self):
        """全选所有表"""
        self.table_list.select_set(0, tk.END)
    
    def deselect_all_tables(self):
        """取消全选所有表"""
        self.table_list.selection_clear(0, tk.END)
    
    def start_export(self):
        """开始导出过程"""
        # 检查数据库连接
        if not self.db_path:
            messagebox.showerror("错误", "请先连接到数据库")
            return
        
        # 获取选中的表
        selected_indices = self.table_list.curselection()
        if not selected_indices:
            messagebox.showerror("错误", "请选择要导出的数据表")
            return
        
        self.selected_tables = [self.table_list.get(i) for i in selected_indices]
        self.export_path = self.export_path_var.get()
        self.export_filename = self.export_filename_var.get().strip() or "export"
        self.export_format = self.export_format_var.get() or "xlsx"
        
        # 检查导出路径
        if not self.export_path:
            messagebox.showerror("错误", "请选择导出路径")
            return
        
        if not os.path.exists(self.export_path):
            messagebox.showerror("错误", "导出路径不存在")
            return
        
        full_export_path = ""
        if self.export_format == "xlsx":
            full_export_path = os.path.join(self.export_path, f"{self.export_filename}.xlsx")
            if os.path.exists(full_export_path):
                if not messagebox.askyesno("确认", "文件已存在，是否覆盖？"):
                    return
        else:
            existing_files = [
                os.path.join(self.export_path, f"{table_name}.csv")
                for table_name in self.selected_tables
                if os.path.exists(os.path.join(self.export_path, f"{table_name}.csv"))
            ]
            if existing_files:
                if not messagebox.askyesno("确认", f"已有 {len(existing_files)} 个同名CSV文件，是否覆盖？"):
                    return
        
        # 准备导出
        self.is_exporting = True
        self.export_cancelled = False
        self.log_messages = []
        self.progress_var.set(0)
        self.progress_label.config(text="准备导出...")
        self.export_button.config(state=tk.DISABLED)
        self.cancel_button.config(state=tk.NORMAL)
        
        # 清空日志
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)
        
        # 在新线程中执行导出
        export_thread = threading.Thread(target=self.export_tables, args=(full_export_path,))
        export_thread.daemon = True
        export_thread.start()
        
        # 定期检查导出状态
        self.root.after(100, self.check_export_status)
    
    def export_tables(self, full_export_path):
        """导出选中的表到文件"""
        try:
            workbook = None
            if self.export_format == "xlsx":
                workbook = openpyxl.Workbook()
                if 'Sheet' in workbook.sheetnames:
                    workbook.remove(workbook['Sheet'])

            encoding = self.get_db_encoding()
            total_tables = len(self.selected_tables)
            for i, table_name in enumerate(self.selected_tables):
                if self.export_cancelled:
                    self.log("导出操作已取消")
                    break
                
                # 更新进度
                progress = (i / total_tables) * 100
                self.progress_var.set(progress)
                self.progress_label.config(text=f"导出表 {table_name} ({i+1}/{total_tables})...")
                
                try:
                    # 连接数据库并导出表
                    conn = sqlite3.connect(self.db_path)
                    cursor = conn.cursor()
                    
                    # 获取表结构
                    cursor.execute(f"PRAGMA table_info('{table_name}');")
                    columns = cursor.fetchall()
                    column_names = [col[1] for col in columns]
                    
                    # 获取表数据
                    cursor.execute(f"SELECT * FROM '{table_name}';")
                    rows = cursor.fetchall()
                    
                    conn.close()
                    
                    if self.export_format == "xlsx":
                        sheet = workbook.create_sheet(title=table_name)
                        for col_idx, col_name in enumerate(column_names, 1):
                            sheet.cell(row=1, column=col_idx, value=col_name)
                        for row_idx, row in enumerate(rows, 2):
                            for col_idx, value in enumerate(row, 1):
                                sheet.cell(row=row_idx, column=col_idx, value=value)
                        for col_idx, col_name in enumerate(column_names, 1):
                            max_width = len(col_name)
                            for row_idx in range(2, len(rows) + 2):
                                cell_value = str(sheet.cell(row=row_idx, column=col_idx).value)
                                if len(cell_value) > max_width:
                                    max_width = len(cell_value)
                            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 2, 50)
                    else:
                        csv_path = os.path.join(self.export_path, f"{table_name}.csv")
                        with open(csv_path, "w", encoding=encoding, newline="") as csv_file:
                            writer = csv.writer(csv_file)
                            writer.writerow(column_names)
                            writer.writerows(rows)
                    
                    self.log(f"成功导出表: {table_name} ({len(rows)} 行数据)")
                except Exception as e:
                    self.log(f"导出表 {table_name} 失败: {e}")
            
            if not self.export_cancelled:
                if self.export_format == "xlsx" and workbook is not None:
                    workbook.save(full_export_path)
                    workbook.close()
                
                self.progress_var.set(100)
                self.progress_label.config(text="导出完成")
                
                if self.export_format == "xlsx":
                    self.log(f"成功导出 {len(self.selected_tables)} 个表到文件: {full_export_path}")
                else:
                    self.log(f"成功导出 {len(self.selected_tables)} 个表到目录: {self.export_path} (编码: {encoding})")
                
                self.root.after(100, lambda: self.export_complete(full_export_path, encoding))
        except Exception as e:
            self.log(f"导出过程中发生错误: {e}")
            self.root.after(100, lambda: messagebox.showerror("错误", f"导出过程中发生错误: {e}"))
        finally:
            self.is_exporting = False
            self.root.after(100, lambda: self.cancel_button.config(state=tk.DISABLED))
            self.root.after(100, lambda: self.export_button.config(state=tk.NORMAL))
    
    def cancel_export(self):
        """取消导出操作"""
        self.export_cancelled = True
        self.progress_label.config(text="取消导出...")
    
    def check_export_status(self):
        """检查导出状态"""
        if self.is_exporting:
            self.root.after(100, self.check_export_status)
    
    def get_db_encoding(self):
        """获取数据库编码并映射到Python编码名"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("PRAGMA encoding")
            row = cursor.fetchone()
            conn.close()
            encoding_name = str((row or ["UTF-8"])[0] or "UTF-8").upper()
        except Exception:
            encoding_name = "UTF-8"
        return {
            "UTF-8": "utf-8",
            "UTF-16": "utf-16",
            "UTF-16LE": "utf-16-le",
            "UTF-16BE": "utf-16-be",
        }.get(encoding_name, "utf-8")

    def export_complete(self, full_export_path, encoding):
        """导出完成后的处理"""
        if self.export_format == "xlsx":
            response = messagebox.askyesnocancel(
                "导出完成", 
                f"成功导出 {len(self.selected_tables)} 个表到文件:\n{full_export_path}\n\n是否打开导出文件？"
            )
            if response is True:
                try:
                    os.startfile(full_export_path)
                except Exception as e:
                    self.log(f"打开文件失败: {e}")
            elif response is False:
                try:
                    os.startfile(os.path.dirname(full_export_path))
                except Exception as e:
                    self.log(f"打开目录失败: {e}")
        else:
            response = messagebox.askyesno(
                "导出完成",
                f"成功导出 {len(self.selected_tables)} 个表到目录:\n{self.export_path}\n编码: {encoding}\n\n是否打开导出目录？"
            )
            if response:
                try:
                    os.startfile(self.export_path)
                except Exception as e:
                    self.log(f"打开目录失败: {e}")
    
    def log(self, message):
        """记录日志信息"""
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        log_message = f"[{timestamp}] {message}"
        self.log_messages.append(log_message)
        
        # 更新日志显示
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, log_message + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)

if __name__ == "__main__":
    root = tk.Tk()
    default_format = "xlsx"
    if len(sys.argv) > 1 and str(sys.argv[1]).lower() in ("xlsx", "csv"):
        default_format = str(sys.argv[1]).lower()
    app = DBExporter(root, default_format=default_format)
    root.mainloop()
